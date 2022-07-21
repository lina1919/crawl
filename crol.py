from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
import datetime
def no_age_no_pic(d,num,k):
    try:
        date = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[1]/ul/li[3]')
        dt = datetime.datetime.strptime(date.text,'%Y-%m-%d')
        if dt>second_date or dt<first_date :
            return 0
        star = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[1]/ul/li[4]/div/span/span')
        s = star.get_attribute("class");
        height = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[1]')
        body = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[2]')
        usual_size = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[3]')
        color = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[2]/span[1]')
        size = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[2]/span[2]')
        review = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[3]/p')
        excel_ws.cell(row = num, column = 1).value = url
        excel_ws.cell(row = num, column = 2).value = cloth
        excel_ws.cell(row = num, column = 3).value = cloth_name.text
        excel_ws.cell(row = num, column = 4).value = date.text
        excel_ws.cell(row = num, column = 5).value = s[-1]
        excel_ws.cell(row = num, column = 7).value=height.text
        excel_ws.cell(row = num, column = 8).value=body.text
        excel_ws.cell(row = num, column = 9).value=usual_size.text
        excel_ws.cell(row = num, column = 10).value=color.text
        excel_ws.cell(row = num, column = 11).value=size.text
        excel_ws.cell(row = num, column = 12).value=review.text
    except:
        print("예외처리")

def yes_age_no_pic(d,num,k):
    try:
        date = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[1]/ul/li[3]')
        dt = datetime.datetime.strptime(date.text,'%Y-%m-%d')
        star = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[1]/ul/li[4]/div/span/span')
        s = star.get_attribute("class");
        if dt>second_date or dt<first_date :
            return 0
        age = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[1]')
        height = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[2]')
        body = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[3]')
        usual_size = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[4]')
        color = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[2]/span[1]')
        size = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[2]/span[2]')
        review = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[3]/p')
        excel_ws.cell(row = num, column = 1).value = url
        excel_ws.cell(row = num, column = 2).value = cloth
        excel_ws.cell(row = num, column = 3).value = cloth_name.text
        excel_ws.cell(row = num, column = 4).value = date.text
        excel_ws.cell(row = num, column = 5).value = s[-1]
        excel_ws.cell(row = num, column = 6).value = age.text
        excel_ws.cell(row = num, column = 7).value=height.text
        excel_ws.cell(row = num, column = 8).value=body.text
        excel_ws.cell(row = num, column = 9).value=usual_size.text
        excel_ws.cell(row = num, column = 10).value=color.text
        excel_ws.cell(row = num, column = 11).value=size.text
        excel_ws.cell(row = num, column = 12).value=review.text
        time.sleep(1)
    except:
        print("예외처리")

def no_age_yes_pic(d,num,k):
    try:
        date = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[1]/ul/li[3]')
        dt = datetime.datetime.strptime(date.text,'%Y-%m-%d')
        if dt>second_date or dt<first_date :
            return 0
        star = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[1]/ul/li[4]/div/span/span')
        s = star.get_attribute("class");
        height = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[1]')
        body = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[2]')
        usual_size = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[3]')
        color = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[2]/span[1]')
        size = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[2]/span[2]')
        review = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[5]/p')
        excel_ws.cell(row = num, column = 1).value = url
        excel_ws.cell(row = num, column = 2).value = cloth
        excel_ws.cell(row = num, column = 3).value = cloth_name.text
        excel_ws.cell(row = num, column = 4).value = date.text
        excel_ws.cell(row = num, column = 5).value = s[-1]
        excel_ws.cell(row = num, column = 7).value=height.text
        excel_ws.cell(row = num, column = 8).value=body.text
        excel_ws.cell(row = num, column = 9).value=usual_size.text
        excel_ws.cell(row = num, column = 10).value=color.text
        excel_ws.cell(row = num, column = 11).value=size.text
        excel_ws.cell(row = num, column = 12).value=review.text
    except:
        print("예외처리")

def yes_age_yes_pic(d,num,k):
    try:
        date = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[1]/ul/li[3]')
        dt = datetime.datetime.strptime(date.text,'%Y-%m-%d')
        if dt>second_date or dt<first_date :
            return 0
        age = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[1]')
        star = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[1]/ul/li[4]/div/span/span')
        s = star.get_attribute("class");
        age = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[1]')
        height = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[2]')
        body = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[3]')
        usual_size = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[1]/span[4]')
        color = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[2]/span[1]')
        size = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[2]/ul/li[2]/span[2]')
        review = driver.find_element(By.XPATH,'//*[@id="evaluation_view'+str(d)+'"]/div[5]/p')
        excel_ws.cell(row = num, column = 1).value = url
        excel_ws.cell(row = num, column = 2).value = cloth
        excel_ws.cell(row = num, column = 3).value = cloth_name.text
        excel_ws.cell(row = num, column = 4).value = date.text
        excel_ws.cell(row = num, column = 5).value = s[-1]
        excel_ws.cell(row = num, column = 6).value = age.text
        excel_ws.cell(row = num, column = 7).value=height.text
        excel_ws.cell(row = num, column = 8).value=body.text
        excel_ws.cell(row = num, column = 9).value=usual_size.text
        excel_ws.cell(row = num, column = 10).value=color.text
        excel_ws.cell(row = num, column = 11).value=size.text
        excel_ws.cell(row = num, column = 12).value=review.text
        time.sleep(1)
    except:
        print("예외처리")
def click_page_num(i):
    try:
        driver.find_element(By.XPATH,'//*[@id="bodyWrap"]/div[2]/div[2]/span/a['+str(i)+']').click()
    except:
        print("예외처리")
def click_review():
    try:
        driver.find_element(By.XPATH,'//*[@id="customerReview"]/a').click()
        time.sleep(1)
    except:
        print("예외처리")
def click_review_num(i):
    try:
        driver.find_element(By.XPATH, '//*[@id="reviewPagingDiv"]/span/a['+str(i)+']').click()
    except:
        print("예외처리")
def click_title(i):
    try:
        a = driver.find_element(By.XPATH,'//*[@id="listBody"]/li['+str(i)+']/div/a[2]/span[2]')
        a.click()
        time.sleep(2)
    except:
        print("예외처리")
def click_review_num(i):
    try:
        driver.find_element(By.XPATH, '//*[@id="reviewPagingDiv"]/span/a['+str(i)+']').click()
        time.sleep(2)
    except:
        print("예외처리")
def close_review_andgoback():
    try:
        driver.find_element(By.XPATH, '//*[@id="customerReviewDiv"]/a').click()
        time.sleep(1)
        driver.back()
        time.sleep(1)
        driver.back()
    except:
        print("예외처리")
driver = webdriver.Chrome()
wb = openpyxl.load_workbook('Time_review_raw data.xlsx')
excel_ws = wb['Sheet2']
num = 61
cloth = '가디건/베스트'
url = 'http://www.thehandsome.com/ko/c/ou_br01_we015/ou#2_0_0_0_0_73_0_0_0'
driver.get(url)
d=0
first_date = datetime.datetime(2021, 6, 1)
second_date = datetime.datetime(2022, 6, 30)
for p in range(1,10):
    for i in range(1,13):
        print('반복문'+str(i)+'번 실행')
        time.sleep(1)
        click_title(i)
        time.sleep(2)
        url = driver.current_url
        cloth_name = driver.find_element(By.XPATH,'//*[@id="contentDiv"]/div[1]/div[1]/h4/span')
        review_num = driver.find_element(By.XPATH,'//*[@id="customerReview"]/a')
        r_num=int(review_num.text[review_num.text.find('(')+1:review_num.text.find(')')])
        page_num = r_num // 4 + 1
        

        print(cloth_name.text)
        print(r_num)

        if r_num == 0:
            print('review 없음')
            driver.back()
            wb.save('Time_review_raw data.xlsx')
        elif r_num <= 4:
            click_review()
            time.sleep(2)
            print('review 개수 4 이하')
            
            for k in range(0,r_num % 4):
                no_age_no_pic(d+k,num+k,5)
                no_age_yes_pic(d+k,num+k,5)
                yes_age_no_pic(d+k,num+k,5)
                yes_age_yes_pic(d+k,num+k,5)
                print('완료')
                time.sleep(2)
            close_review_andgoback()
            num += r_num%4
            wb.save('Time_review_raw data.xlsx')
        else:
            click_review()
            time.sleep(2)
            print('review 개수 5 이상')
            for j in range(0,page_num-1):
                if (j+2 > 10):
                    print("리뷰 너무 많음")
                else:
                    for k in range(0,4):
                        no_age_no_pic(d+k,num+k,5)
                        no_age_yes_pic(d+k,num+k,5)
                        yes_age_no_pic(d+k,num+k,5)
                        yes_age_yes_pic(d+k,num+k,5)
                        print('완료')
                    click_review_num(j+2)

                num += 4
            for k in range(0,r_num % 4):
                no_age_no_pic(d+k,num+k,5)
                no_age_yes_pic(d+k,num+k,5)
                yes_age_no_pic(d+k,num+k,5)
                yes_age_yes_pic(d+k,num+k,5)
                print('완료')
            num += r_num%4
            print('완료')
            close_review_andgoback()
            wb.save('Time_review_raw data.xlsx')
    click_page_num(p+1)
wb.save('Time_review_raw data.xlsx')



